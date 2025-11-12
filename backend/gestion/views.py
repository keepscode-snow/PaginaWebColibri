from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Producto, Venta, DetalleVenta, Pedido
from .decorators import cajero_required
from django.contrib import messages
from decimal import Decimal
from .decorators import admin_required, cajero_required
from django.db import transaction
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django import forms
from django.db.models import Sum
from django.utils.timezone import make_aware
from datetime import time




@login_required
def home_redirect(request):
    if request.user.rol == 'admin':
        return redirect('dashboard_admin')
    elif request.user.rol == 'cajero':
        return redirect('registrar_venta')  # O 'dashboard_cajero' si prefieres
    else:
        return redirect('login')  # fallback


@login_required
def menu_inicio(request):
    user = request.user
    es_cajero = user.groups.filter(name='Cajero').exists()
    es_admin = user.groups.filter(name='Administrador').exists()

    return render(request, 'gestion/menu_inicio.html', {
        'es_cajero': es_cajero,
        'es_admin': es_admin
    })


@cajero_required
def registrar_venta(request):
    if 'carrito' not in request.session:
        request.session['carrito'] = {}

    carrito = request.session['carrito']
    productos = []

    for sku, item in carrito.items():
        producto = Producto.objects.get(sku=sku)
        productos.append({
            'producto': producto,
            'cantidad': item['cantidad'],
            'subtotal': producto.precio * item['cantidad']
        })

    total = sum([p['subtotal'] for p in productos])

    # Obtener número de boleta sugerido
    ultima_venta = Venta.objects.order_by('-numero_boleta').first()
    numero_siguiente = ultima_venta.numero_boleta + 1 if ultima_venta else 1

    return render(request, 'gestion/venta.html', {
        'productos': productos,
        'total': total,
        'numero_sugerido': numero_siguiente
    })

@cajero_required
def agregar_al_carrito(request):
    if request.method == 'POST':
        busqueda = request.POST['busqueda']
        cantidad = int(request.POST['cantidad'])

        try:
            producto = Producto.objects.get(sku=busqueda)
        except Producto.DoesNotExist:
            producto = Producto.objects.filter(nombre__icontains=busqueda).first()

        if producto:
            carrito = request.session.get('carrito', {})
            sku = str(producto.sku)

            if sku in carrito:
                carrito[sku]['cantidad'] += cantidad
            else:
                carrito[sku] = {'cantidad': cantidad}

            request.session['carrito'] = carrito
            messages.success(request, f'{producto.nombre} agregado al carrito.')
        else:
            messages.error(request, 'Producto no encontrado.')

    return redirect('registrar_venta')

@cajero_required
@transaction.atomic
def confirmar_venta(request):
    if request.method == 'POST':
        carrito = request.session.get('carrito', {})
        if not carrito:
            messages.error(request, "El carrito está vacío.")
            return redirect('registrar_venta')

        # Generar automáticamente el próximo número de boleta
        ultimo = Venta.objects.order_by('-numero_boleta').first()
        nuevo_numero_boleta = ultimo.numero_boleta + 1 if ultimo else 1

        if not nuevo_numero_boleta:
            messages.error(request, "Debe ingresar un número de boleta.")
            return redirect('registrar_venta')

        if Venta.objects.filter(numero_boleta=nuevo_numero_boleta).exists():
            messages.error(request, "Ese número de boleta ya fue registrado.")
            return redirect('registrar_venta')

        venta = Venta.objects.create(
            numero_boleta=nuevo_numero_boleta,
            total=0,
            medio_pago='Efectivo',
            creado_por=request.user,
            fecha=timezone.now()
        )

        total_venta = 0

        for sku, item in carrito.items():
            producto = Producto.objects.get(sku=sku)
            cantidad = item['cantidad']

            if producto.stock < cantidad:
                messages.error(request, f"Stock insuficiente para {producto.nombre}.")
                venta.delete()
                return redirect('registrar_venta')

            subtotal = producto.precio * cantidad
            total_venta += subtotal

            DetalleVenta.objects.create(
                venta=venta,
                producto=producto,
                cantidad=cantidad,
                precio_unitario=producto.precio
            )

            producto.stock -= cantidad
            producto.save()

        venta.total = total_venta
        venta.save()

        request.session['carrito'] = {}  # Vaciar carrito
        messages.success(request, f"✅ Venta registrada correctamente. Boleta N°{venta.numero_boleta} - Total: ${total_venta}")
        return redirect('registrar_venta')

    return redirect('registrar_venta')



@cajero_required
def registrar_pedido(request):
    if request.method == 'POST':
        cliente_nombre = request.POST.get('cliente_nombre')
        cliente_telefono = request.POST.get('cliente_telefono')
        fecha_entrega = request.POST.get('fecha_entrega')
        anticipo = request.POST.get('anticipo')
        total = request.POST.get('total')

        if not all([cliente_nombre, cliente_telefono, fecha_entrega, anticipo, total]):
            messages.error(request, "Todos los campos son obligatorios.")
            return redirect('registrar_pedido')

        pedido = Pedido.objects.create(
            cliente_nombre=cliente_nombre,
            cliente_telefono=cliente_telefono,
            fecha_entrega=fecha_entrega,
            anticipo=anticipo,
            total=total,
            estado='BORRADOR',  # o 'PENDIENTE'
            creado_por=request.user
        )

        messages.success(request, f"Pedido #{pedido.id} registrado correctamente.")
        return redirect('registrar_pedido')

    return render(request, 'gestion/registrar_pedido.html')

@cajero_required
def lista_pedidos(request):
    pedidos = Pedido.objects.filter(creado_por=request.user).order_by('-fecha_entrega')
    return render(request, 'gestion/lista_pedidos.html', {'pedidos': pedidos})

@cajero_required
def cambiar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, creado_por=request.user)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        if nuevo_estado in dict(Pedido.ESTADOS).keys():
            pedido.estado = nuevo_estado
            pedido.save()
            messages.success(request, f"Estado del pedido #{pedido.id} actualizado a {nuevo_estado}.")
        else:
            messages.error(request, "Estado inválido.")

    return redirect('lista_pedidos')



@staff_member_required
def lista_pedidos_admin(request):
    pedidos = Pedido.objects.all().order_by('-fecha_entrega')

    estado = request.GET.get('estado')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if estado:
        pedidos = pedidos.filter(estado=estado)

    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_entrega__gte=fecha_inicio_dt)
        except ValueError:
            messages.error(request, 'Formato de fecha de inicio inválido')

    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_entrega__lte=fecha_fin_dt)
        except ValueError:
            messages.error(request, 'Formato de fecha de fin inválido')

    return render(request, 'gestion/lista_pedidos_admin.html', {
        'pedidos': pedidos,
        'estado_seleccionado': estado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })

@staff_member_required
def admin_cambiar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        if nuevo_estado in dict(Pedido.ESTADOS).keys():
            pedido.estado = nuevo_estado
            pedido.save()
            messages.success(request, f"Estado del pedido #{pedido.id} cambiado a {nuevo_estado}.")
        else:
            messages.error(request, "Estado inválido.")
    return redirect('lista_pedidos_admin')


@staff_member_required
def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.all().order_by('-fecha_entrega')

    estado = request.GET.get('estado')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if estado:
        pedidos = pedidos.filter(estado=estado)

    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_entrega__gte=fecha_inicio_dt)
        except ValueError:
            pass

    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_entrega__lte=fecha_fin_dt)
        except ValueError:
            pass

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Encabezados
    encabezados = ['ID', 'Cliente', 'Teléfono', 'Fecha Entrega', 'Anticipo', 'Total', 'Estado', 'Cajero']
    ws.append(encabezados)

    for pedido in pedidos:
        ws.append([
            pedido.id,
            pedido.cliente_nombre,
            pedido.cliente_telefono,
            pedido.fecha_entrega.strftime('%Y-%m-%d %H:%M'),
            float(pedido.anticipo),
            float(pedido.total),
            pedido.estado,
            pedido.creado_por.username
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(encabezados, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    wb.save(response)
    return response



@staff_member_required
def lista_productos_admin(request):
    productos = Producto.objects.all().order_by('nombre')
    return render(request, 'gestion/productos/lista.html', {'productos': productos})

class ProductoForm(forms.ModelForm):
    class Meta:
        model = Producto
        fields = ['sku', 'nombre', 'precio', 'stock', 'activo', 'categoria']

@staff_member_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('lista_productos_admin')
    else:
        form = ProductoForm()
    return render(request, 'gestion/productos/formulario.html', {'form': form, 'accion': 'Crear'})


@staff_member_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado.')
            return redirect('lista_productos_admin')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'gestion/productos/formulario.html', {'form': form, 'accion': 'Editar'})

@staff_member_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    producto.delete()
    messages.success(request, 'Producto eliminado.')
    return redirect('lista_productos_admin')




@admin_required
def dashboard_admin(request):
    hoy = timezone.localdate()
    inicio_dia = make_aware(datetime.combine(hoy, time.min))
    fin_dia = make_aware(datetime.combine(hoy, time.max))

    ventas_hoy = Venta.objects.filter(fecha__range=(inicio_dia, fin_dia))
    
    total_ventas_hoy = ventas_hoy.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas_hoy = ventas_hoy.count()

    productos_vendidos = DetalleVenta.objects.filter(
        venta__fecha__date=hoy
    ).aggregate(total=Sum('cantidad'))['total'] or 0

    pedidos_pendientes = Pedido.objects.filter(estado__in=['BORRADOR', 'CONFIRMADO']).count()

    return render(request, 'gestion/dashboard_admin.html', {
        'total_ventas_hoy': total_ventas_hoy,
        'cantidad_ventas_hoy': cantidad_ventas_hoy,
        'productos_vendidos': productos_vendidos,
        'pedidos_pendientes': pedidos_pendientes
    })


@cajero_required
def dashboard_cajero(request):
    return render(request, 'gestion/dashboard_cajero.html')
